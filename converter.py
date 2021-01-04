# -*- coding: utf-8 -*-
import sys
import re
import os
import openpyxl 
from openpyxl import Workbook
from time import sleep
from progress.spinner import MoonSpinner
import threading
import itertools

def run_progress_bar(finished_event):
    with MoonSpinner('Processing...') as bar:
        while not finished_event.is_set():
            sleep(0.02)
            bar.next()
            finished_event.wait(0.2)
        
if len(sys.argv) != 1:
    print("Usage: python3 program.py")
    exit(1)

oldLayout = False
newLayout = False
for filename in os.listdir():
    if filename.endswith(".html"):
        print("File: {}".format(filename))
        with open(filename, 'r') as f:
            contents = f.read()
            if "<html lang=\"en\" id=\"facebook\"" in contents:
                oldLayout = True
            elif "<html id=\"facebook\"" in contents:
                newLayout = True
            else:
                print("Format Error.")
                exit(1)
        finished_event = threading.Event()
        progress_bar_thread = threading.Thread(target=run_progress_bar, args=(finished_event,))
        progress_bar_thread.start()
        
        if oldLayout:
            with open(filename, 'r') as f:
                contents = f.read()
                start = contents.find("data-ft=\"{&quot;tn&quot;:&quot;K&quot;}\" id=\"js_") + len("data-ft=\"{&quot;tn&quot;:&quot;K&quot;}\" id=\"js_")
                end = contents.find("About</a>")
                contents = contents[start-48:end]
                contents = re.sub('<br><br>', ' ', contents)
                contents = re.sub('<br>', ' ', contents)
                contents = re.sub('<div class="_2fj2"><div class="_3mvw', '\n<div class="_2fj2"><div class="_3mvw', contents)                 
                contents = re.sub('</div></div><div class="_3mvr"', '</div></div><div class="_3mvr"\n', contents)                 
                contents = re.sub('<li><div><div aria-label="', '\n', contents) 
                contents = re.sub('data-testid="post_message"', '\n', contents)
                contents = re.sub('<div class="_6ks">', '\n<div class="_6ks">', contents)
            text_file = open("output1.txt", "w")
            n = text_file.write(contents)
            text_file.close()     
                
            outfile = open("output2.txt", "w+")
            for line in open("output1.txt", "r"):
                if "Comment" in line or "data-ft=\"{&quot;tn&quot;:&quot;K&quot;}\" id=\"js_" in line:
                    if "</div></div><div class=\"_3mvr\"" not in line:
                        if "<div class=\"_6ks\">" not in line:
                            if "\" role=\"article\" class=\"_4eek" in line and "[object Object]\" direction=\"left\">" in line:
                                firstDelPos = line.find("\" role=\"article\" class=\"_4eek") + len("\" role=\"article\" class=\"_4eek")
                                secondDelPos = line.find("[object Object]\" direction=\"left\">") + len("[object Object]\" direction=\"left\">")
                                line = line.replace(line[firstDelPos-29:secondDelPos], "")
                            outfile.write(line)
            outfile.close()
            
            with open('output2.txt', 'r') as f:        
                contents = f.read()
                contents = re.sub('<span class="_81hb">.*?</span>', '', contents)
                contents = re.sub('<span class="_1lld">.*?</span>', '', contents)
                contents = re.sub('<[^<]+?>', '', contents)
                contents = re.sub('&nbsp;', '', contents)
                contents = re.sub('&amp;', '&', contents)
                contents = re.sub('Author', '', contents)
                contents = re.sub('Write a reply…', '', contents)
                contents = re.sub('Write a comment…', '', contents)
                years = 1
                for years in range(10):
                    contents = re.sub("·Like·Reply·{}y".format(years),"", contents) # years
                for weeks in range(53):
                    contents = re.sub("·Like·Reply·{}w".format(weeks),"", contents) # weeks

                days = 1
                for days in range(7):
                    contents = re.sub("·Like·Reply·{}d".format(days),"", contents) # days
                contents = re.sub('Manage', '', contents)
                contents = re.sub('·Edited','', contents)
                contents = re.sub('Edited','', contents)
                contents = re.sub('&lt;','<', contents)
                contents = re.sub('data-ft="{&quot;tn&quot;:&quot;K&quot;}" id="js_a">' , '', contents)
                contents = re.sub('Active Now','', contents)
            text_file = open("output3.txt", "w")
            n = text_file.write(contents)
            text_file.close()     

            file1 = open('output3.txt', 'r') 
            book = Workbook()
            sheet = book.active

            count = 0
            while True: 
                count += 1
                line = file1.readline() 
                if not line: 
                    break
                line = re.sub('\n','', line)
                sheet.cell(row=count, column=1).value = line
                book.save('output1test.xlsx')
            
        if newLayout:
            with open(filename, 'r') as f:
                contents = f.read()
                contents = re.sub('class="a8c37x1j ni8dbmo4 stjgntxs l9j0dhe7 ltmttdrg g0qnabr5 ojkyduve">youtube.com</span></span></div><div', '', contents)
                start = contents.find("Subject: ") + len("Subject: ")
                end = contents.find("Write a comment…")
                contents = contents[start-9:end]
                contents = re.sub('Comment by', '\nComment just', contents) 
                contents = re.sub('Reply by', '\nComment reply', contents)
                contents = re.sub('class="gpro0wi8 pcp91wgn">', '\nclass="gpro0wi8 pcp91wgn">', contents)
                contents = re.sub('<span dir="auto">', '\n<span dir="auto">', contents)
                contents = re.sub('&nbsp;·', '', contents)
                contents = re.sub('&nbsp;', '', contents)
                
                years = 1
                for years in range(10):
                    contents = re.sub("{} years".format(years),"", contents) # years
                for weeks in range(53):
                    contents = re.sub("{} weeks".format(weeks),"", contents) # weeks

                days = 1
                for days in range(7):
                    contents = re.sub("{} days".format(days),"", contents) # days
                    
            text_file = open("output1.txt", "w")
            n = text_file.write(contents)
            text_file.close()      
                
            outfile = open("output2.txt", "w+")
            for line in open("output1.txt", "r"):
                line = re.sub('ago" role="article" tabindex="-1" class="l9j0dhe7 ecm0bbzt rz4wbd8a qt6c0cv9 dati1w0a j83agx80 btwxx1t3 lzcic4wl">', '', line)
                line = re.sub('</div></div></span></div></div></div></div><div class="_6cuq', '\n</div></div></span></div></div></div></div><div class="_6cuq', line)
                line = re.sub('ago" role="article" tabindex="-1" class="l9j0dhe7 ecm0bbzt rz4wbd8a qt6c0cv9 scb9dxdr j83agx80 btwxx1t3 lzcic4wl">', '', line)
                outfile.write(line)
            outfile.close()

            outfile = open("output3.txt", "w+")
            for line in open("output2.txt", "r"):
                if "</div></div></span></div></div></div></div><div class=\"_6cuq" not in line:
                    if "class=\"gpro0wi8 pcp91wgn\">" not in line:
                        if "<span dir=\"auto\">" not in line:
                            if "Comment just" in line and "lrazzd5p oo9gr5id\" dir=\"auto\">" in line:
                                firstDelPos= line.find("Comment just")
                                secondDelPos = line.find("lrazzd5p oo9gr5id\" dir=\"auto\">")
                                line = line.replace(line[firstDelPos+7:secondDelPos+30], "")
                            if "Comment reply" in line and "lrazzd5p oo9gr5id\" dir=\"auto\">" in line:
                                firstDelPos= line.find("Comment reply")
                                secondDelPos = line.find("lrazzd5p oo9gr5id\" dir=\"auto\">")
                                line = line.replace(line[firstDelPos+13:secondDelPos+30], "")                
                            line = re.sub('<[^<]+?>', '', line)
                            outfile.write(line)
            outfile.close()


            with open("output3.txt", 'r') as f:
                
                contents = f.read()
                contents = re.sub('<div aria-label="', '', contents)
                contents = re.sub('Edited', '', contents)
                
                like = 1
                for like in reversed(range(200)) :
                    years = 1
                    for years in range(10):
                        contents = re.sub("{}LikeReply{}y".format(like,years),"", contents) # years
                        contents = re.sub("LikeReply{}y".format(years),"", contents) # years
                    weeks = 1
                    for weeks in range(53):
                        contents = re.sub("{}LikeReply{}w".format(like,weeks),"", contents) # weeks
                        contents = re.sub("LikeReply{}w".format(weeks),"", contents) # weeks
                    days = 1
                    for days in range(7):
                        contents = re.sub("{}LikeReply{}d".format(like, days),"", contents) # days
                        contents = re.sub("LikeReply{}d".format(days),"", contents) # days
                contents = re.sub("Write a reply…","", contents)  
                    
            text_file = open("output4.txt", "w")
            n = text_file.write(contents)
            text_file.close() 

            file1 = open('output4.txt', 'r') 
            book = Workbook()
            sheet = book.active

            count = 0
            while True: 
                count += 1
                line = file1.readline() 
                if not line: 
                    break
                line = re.sub('\n','', line)
                sheet.cell(row=count, column=1).value = line
                book.save('output1test.xlsx')
            os.system("rm output4.txt")
        finished_event.set()
        progress_bar_thread.join()
        os.system("rm output1.txt")
        os.system("rm output2.txt")
        os.system("rm output3.txt")        
        finished_event = threading.Event()
        progress_bar_thread = threading.Thread(target=run_progress_bar, args=(finished_event,))
        progress_bar_thread.start()        
        if oldLayout:
            try:
                contents = ""
                with open(filename, 'r') as f:
                    contents = f.read()
                # matches = re.findall('"><img alt=".+?" class="',contents)
                    matches = re.sub("\">", '\n', contents)
                #matches = re.findall('">.+?</a> <span dir="ltr"><span class="',contents)
                # print(matches)
                text_file = open("A.txt", "w+")
                n = text_file.write(matches)
                text_file.close()

                good_words = ['</a> <span dir="', '</a><span class="_2ltv']
                outfile = open("nodup.txt", "w+")
                for line in open("A.txt", "r"):
                    if any(good_words in line for good_words in good_words):
                        outfile.write(line)
                outfile.close()

                outfile = open("B.txt", "w+")
                lines_seen = set()  # holds lines already seen
                for line in open("nodup.txt", "r"):
                    if line not in lines_seen:  # not a duplicate
                        outfile.write(line)
                        lines_seen.add(line)
                outfile.close()

                with open("B.txt", 'r') as f:
                    contents = f.read()
                    matches = re.sub("</a> <span dir=\"ltr", '', contents)
                    matches = re.sub("</a><span class=\"_2ltv", '', matches)
                text_file = open("C.txt", "w+")
                n = text_file.write(matches)
                text_file.close()

                file1 = open('C.txt', 'r')
                book = Workbook()
                sheet = book.active
                count = 0
                while True:
                    count += 1
                    # Get next line from file
                    line = file1.readline()
                    if not line:
                        break
                    # print(count)
                    line = re.sub('\n', '', line)
                    sheet.cell(row=count, column=1).value = line
                    book.save('Name.xlsx')
                os.system("rm A.txt")
                os.system("rm B.txt")
                os.system("rm C.txt")
                os.system("rm nodup.txt")
                f.close()
            except KeyboardInterrupt:
                os.system("rm A.txt")
                os.system("rm B.txt")
                os.system("rm C.txt")
                os.system("rm nodup.txt")
                os.system("rm Name.xlsx")
    
        if newLayout:
            
            try:
                contents = ""
                with open(filename, 'r') as f:
                    contents = f.read()
                    matches = re.sub("dir=\"auto\">", '\n', contents)
                    
                text_file = open("A.txt", "w+")
                n = text_file.write(matches)
                text_file.close()
                
                lines_seen = set() # holds lines already seen
                good_words = ['</span></span></a>','</span></span></div>']
                bad_words = ['</div></div></div></div></div>', '?', '</', '="', '[']
                outfile = open("nodup.txt", "w+")
                for line in open("A.txt", "r"):
                    if any(good_words in line for good_words in good_words):
                        line = re.sub('</span></span></a>.*?oo9gr5id"', '', line)
                        line = re.sub('</span></span></div>.*?oo9gr5id"', '', line)
                            
                        line = re.sub('</span></span></a></div>.*?m9osqain"', '', line)
                        if not any(bad_word in line for bad_word in bad_words):
                            line = re.sub(' \n', '\n', line)   
                            if line not in lines_seen:             
                                outfile.write(line)
                                lines_seen.add(line)            
                outfile.close()

                file1 = open('nodup.txt', 'r')
                book = Workbook()
                sheet = book.active
                count = 0
                while True:
                    count += 1
                    # Get next line from file
                    line = file1.readline()
                    if not line:
                        break
                    # print(count)
                    line = re.sub('\n', '', line)
                    sheet.cell(row=count, column=1).value = line
                    book.save('Name.xlsx')
                os.system("rm A.txt")
                os.system("rm B.txt")
                os.system("rm C.txt")
                os.system("rm nodup.txt")
                f.close()
            except KeyboardInterrupt:
                os.system("rm A.txt")
                os.system("rm B.txt")
                os.system("rm C.txt")
                os.system("rm nodup.txt")
                os.system("rm Name.xlsx")
        finished_event.set()
        progress_bar_thread.join()
        wb = openpyxl.Workbook()
        wb.save('Destination.xlsx')
        finished_event = threading.Event()
        progress_bar_thread = threading.Thread(target=run_progress_bar, args=(finished_event,))
        progress_bar_thread.start()
        
        book = openpyxl.load_workbook('output1test.xlsx')
        sheet = book.active
        Data_rowsCounts = sheet.max_row

        count = 0
        while True:
            count += 1
            found = 0
            CommentReply = 0
            if count > Data_rowsCounts: 
                break
            book = openpyxl.load_workbook('output1test.xlsx')
            sheet = book.active
            Data_rowsCounts = sheet.max_row
            cell_Data = sheet.cell(row=count, column=1).value
            if cell_Data != None:
                book = openpyxl.load_workbook('Name.xlsx')
                sheet = book.active
                Name_rowsCounts = sheet.max_row
                for j in range(1, Name_rowsCounts + 1):

                    cell_Name = sheet.cell(row=j, column=1).value
                    if "Comment reply" in cell_Data:
                        cell_Data = re.sub("Comment reply","",cell_Data)
                        CommentReply = 1
                    if "Comment" in cell_Data:
                        cell_Data = re.sub("Comment","",cell_Data)        
                    if cell_Name in cell_Data and cell_Name[0] == cell_Data.lstrip(" ")[0]: 
                        book = openpyxl.load_workbook('Destination.xlsx')
                        sheet = book.active
                        sheet.cell(row=count, column=1+CommentReply).value = cell_Name.lstrip(" ")
                        RemainingData = re.sub(cell_Name,"",cell_Data)
                        sheet.cell(row=count, column=1+CommentReply+1).value = RemainingData.lstrip(" ")
                        book.save('Destination.xlsx')
                        found = 1
                        break
            if found == 0:
                book = openpyxl.load_workbook('Destination.xlsx')
                sheet = book.active
                if cell_Data != None:
                    sheet.cell(row=count, column=1).value = cell_Data.lstrip(" ")
                book.save('Destination.xlsx')
            found = 0
            cell_C_Data = sheet.cell(row=count, column=3).value
            
            if cell_C_Data != None:
                for j in range(1, Name_rowsCounts + 1):
                    
                    book = openpyxl.load_workbook('Name.xlsx')
                    sheet = book.active
                    cell_Name = sheet.cell(row=j, column=1).value
                    
                    if cell_Name in cell_C_Data:
                        cell_C_Data = "@" + cell_C_Data
                        book = openpyxl.load_workbook('Destination.xlsx')
                        sheet = book.active
                        sheet.cell(row=count, column=3).value = cell_C_Data
                        book.save('Destination.xlsx')
        finished_event.set()
        progress_bar_thread.join()
        filename = re.sub(' ', r'\ ', filename)
        filename = re.sub('.html', '', filename)
        os.system("cp Destination.xlsx %s.xlsx" %filename)
        os.system("rm Name.xlsx")
        os.system("rm Destination.xlsx")
        os.system("rm output1test.xlsx")
        print("Success.") 
        print('-' * 50) 